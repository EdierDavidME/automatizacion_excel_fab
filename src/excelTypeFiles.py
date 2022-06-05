from openpyxl import Workbook, load_workbook
from src.utils.validations import Validations as vld
from datetime import datetime


class ExcelFiles:
    def __init__(self, file, dir, sheet):
        # self.ChangeSheet =
        self.vld = vld()
        self.data = []
        self.duplicate = []

        self.wb = None
        self.dir = dir
        self.file = file
        self.sheet = sheet
        self.ws = None
        self.sheet_names = None
        self.sheet_name = None

    def write(self):
        self.wb = Workbook()
        self.ws = self.wb.active
        self.ws.title = "Excel Files"
        self.ws.append(["File Name", "File Type"])
        self.ws.append(["test.xlsx", "Excel"])
        self.ws.append(["test.xls", "Excel"])
        self.ws.append(["test.xlsm", "Excel"])
        self.ws.append(["test.xlsb", "Excel"])

        self.wb.save(self.dir + "\\excelTypeFiles.xlsx")

    def read(self):
        self.wb = load_workbook(self.file)
        self.sheet_names = self.wb.sheetnames

        try:
            # GET NITs
            current_row = 2
            dict = None

            for position, row in enumerate(self.wb['test'].iter_rows(min_col=1, min_row=current_row)):
                dict = None
                count_duplicates, current_row, current_row2 = 0, current_row+1, current_row+1

                for index, search in enumerate(self.wb['test'].iter_rows(min_col=1, min_row=current_row)):
                    count = 0
                    
                    if row[0].value not in self.duplicate and row[0].value == search[0].value:
                        self.duplicate.append(row[0].value)
                        count_duplicates += 1
                        if count == 0:
                            dict = {'position': current_row-1,
                                    'cod': row[3].value}

                        if dict['cod'] > search[3].value:
                            dict['position'] = current_row2
                            dict['cod'] = search[3].value
                    current_row2 += 1
                    count += 1

                if count_duplicates == 0:
                    if row[0].value not in self.duplicate:
                        self.wb['test'][f'E{current_row-1}'] = 'Y'
                if dict is not None:
                    self.wb['test'][f'E{dict["position"]}'] = 'Y'
            self.wb.save(self.file)
            
            return datetime.now()

        except Exception as e:
            print(e)
        return self.sheet_names[1]

    # def validateNitExit(self):
    #     for index, value in enumerate(self.data):
    #         if value.

    def get_name_sheet(self):
        try:
            # if self.vld.is_empty(self.sheet) else self.sheet
            self.sheet_name = int(self.sheet)
            return 0
        except Exception as e:
            return self.sheet_names.index(self.sheet)


class ChangeSheet:
    def __init__(self, row, NIT, sitios, nan, COD, isPrincipal, Nombre):
        self.row = row
        self.NIT = NIT
        self.sitios = sitios
        self.nan = nan
        self.COD = COD
        self.isPrincipal = isPrincipal
        self.Nombre = Nombre

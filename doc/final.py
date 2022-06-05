# VALIDATIONS
from openpyxl import Workbook, load_workbook
import os
import re


class Validations:
    def __init__(self):
        pass

    def is_empty(self, value):
        if value == "None" or value == None or value == "" or value == " " or len(re.sub(r"\s+", "", str(value))) == 0:
            return True
        else:
            return False

    def is_file(self, file_path):
        if os.path.isfile(file_path):
            return True
        else:
            return False

    def is_dir(self, file_path):
        if os.path.isdir(file_path):
            return True
        else:
            return False

    def is_number(self, value):
        try:
            int(value)
            return True
        except ValueError:
            return False

    def is_float(self, value):
        try:
            float(value)
            return True
        except ValueError:
            return False

    def is_integer(self, value):
        try:
            int(value)
            return True
        except ValueError:
            return False

    def is_boolean(self, value):
        if value == True or value == False:
            return True
        else:
            return False

    def is_list(self, value):
        if type(value) == list:
            return True
        else:
            return False

    def is_dict(self, value):
        if type(value) == dict:
            return True
        else:
            return False

    def is_string(self, value):
        if type(value) == str:
            return True
        else:
            return False

    def is_tuple(self, value):
        if type(value) == tuple:
            return True
        else:
            return False

    def is_set(self, value):
        if type(value) == set:
            return True
        else:
            return False

    def is_none(self, value):
        if value == None:
            return True
        else:
            return False


# EXCEL FUNCTIONS


class ExcelFiles:
    def __init__(self, file, dir, sheet):
        # self.ChangeSheet =
        self.vld = Validations()
        self.data = []
        self.duplicate = []

        self.wb = None
        self.dir = dir
        self.file = file
        self.sheet = sheet
        self.ws = None
        self.sheet_names = None
        self.sheet_name = None

    def write(self):
        self.wb = Workbook()
        self.ws = self.wb.active
        self.ws.title = "Excel Files"
        self.ws.append(["File Name", "File Type"])
        self.ws.append(["test.xlsx", "Excel"])
        self.ws.append(["test.xls", "Excel"])
        self.ws.append(["test.xlsm", "Excel"])
        self.ws.append(["test.xlsb", "Excel"])

        self.wb.save(self.dir + "\\excelTypeFiles.xlsx")

    def read(self):
        self.wb = load_workbook(self.file)
        self.sheet_names = self.wb.sheetnames

        try:
            # GET NITs
            current_row = 2
            dict = None

            for position, row in enumerate(self.wb['test'].iter_rows(min_col=1, min_row=current_row)):
                dict = None
                count_duplicates, current_row, current_row2 = 0, current_row+1, current_row+1

                for index, search in enumerate(self.wb['test'].iter_rows(min_col=1, min_row=current_row)):
                    count = 0

                    if row[0].value not in self.duplicate and row[0].value == search[0].value:
                        self.duplicate.append(row[0].value)
                        count_duplicates += 1
                        if count == 0:
                            dict = {'position': current_row-1,
                                    'cod': row[3].value}

                        if dict['cod'] > search[3].value:
                            dict['position'] = current_row2
                            dict['cod'] = search[3].value
                    current_row2 += 1
                    count += 1

                if count_duplicates == 0:
                    if row[0].value not in self.duplicate:
                        self.wb['test'][f'E{current_row-1}'] = 'Y'
                if dict is not None:
                    self.wb['test'][f'E{dict["position"]}'] = 'Y'
            self.wb.save(self.file)
            return 'Done!'

        except Exception as e:
            print(e)
        return self.sheet_names[1]

    # def validateNitExit(self):
    #     for index, value in enumerate(self.data):
    #         if value.

    def get_name_sheet(self):
        try:
            # if self.vld.is_empty(self.sheet) else self.sheet
            self.sheet_name = int(self.sheet)
            return 0
        except Exception as e:
            return self.sheet_names.index(self.sheet)


class ChangeSheet:
    def __init__(self, row, NIT, sitios, nan, COD, isPrincipal, Nombre):
        self.row = row
        self.NIT = NIT
        self.sitios = sitios
        self.nan = nan
        self.COD = COD
        self.isPrincipal = isPrincipal
        self.Nombre = Nombre

# READ FILES

class ReadFiles:
    """
    ReadFiles class
    """

    def __init__(self, file_name, folder_ouput, sheet):
        """
        Constructor
        """
        self.sheet = sheet
        self.temp_path = folder_ouput
        self.Validations = Validations()
        self.file_extension = file_name.split(".")[1]
        self.file_name = file_name
        self.folder_path = os.path.dirname(
            os.path.abspath(__file__)) + "\\assets"
        self.folder_ouput_save = folder_ouput if self.Validations.is_dir(self.temp_path) else os.path.dirname(
            os.path.abspath(__file__)) + "\\assets"
        self.path_file = os.path.join(self.folder_path, self.file_name)
        # self.file_content = self.read_file()

    def read_file(self):
        """
        Reads the file
        """

        try:
            print("> Reading file: {}".format(self.file_name.split(".")[0]))
            print("> File path: {}".format(self.path_file))

            if(self.Validations.is_empty(self.file_name)):
                raise Exception("File name is empty")

            elif(self.Validations.is_file(self.path_file)):
                if self.file_extension == "txt":
                    return self.read_txt_file()
                elif self.file_extension == "xlsx":
                    return self.read_excel_file()
            else:
                raise Exception("File does not exist")

        except Exception as e:
            return("* Error: {}".format(e))

    def read_txt_file(self):
        """
        Reads a txt file
        """
        try:
            with open(self.path_file, 'r') as file:
                file_content = file.read()
            # print("> File content: {}".format(file_content))
            return file_content

        except Exception as e:
            return("* Error: {}".format(e))

    def read_excel_file(self):
        """
        Reads an excel file
        """
        try:
            excel = ExcelFiles(self.path_file, self.folder_ouput_save, self.sheet)
            return excel.read()

        except Exception as e:
            return("* Error: {}".format(e))

# MAIN
# PRINCIPAL CLASS: Main
# PURPOSE: Main class for the program.
import argparse as ap
from datetime import datetime

LINES = 25

parser = ap.ArgumentParser()
parser.add_argument("-f", "--file", help="File to be read", required=True)
parser.add_argument("-s", "--sheet", help="Sheet from excel", required=False)
parser.add_argument("-o", "--ouput", help="Dir to save", required=False)
args = parser.parse_args()


# def main():
#     if __name__ == "main":
print("#\tExcel\t\t#")
print("-"*LINES)
print("START: ", datetime.now())
read_file = ReadFiles(args.file, str(args.ouput), str(args.sheet))
print(read_file.read_file())


# main()
